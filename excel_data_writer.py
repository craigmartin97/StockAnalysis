import openpyxl


def write_sp500_stock_information(path, sheet, data):
    if path is None or sheet is None or data is None:
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[sheet]
    max_row = sheet.max_row
    max_col = sheet.max_column
    for stock in data:
        ticker = stock.get_ticker()
        for row in range(3, max_row + 1):
            ticker_obj = sheet.cell(row=row, column=1)

            if ticker_obj.value != ticker:  # Found matching stock
                continue

            financial_records = stock.get_data()
            for financial in financial_records:
                # Store financial info
                name = financial.get_name()
                yearOne = financial.get_year_one()
                yearTwo = financial.get_year_two()
                yearThree = financial.get_year_three()

                total = 3
                if yearOne is None:
                    yearOne = 0
                    total -= 1
                if yearTwo is None:
                    yearTwo = 0
                    total -= 1
                if yearThree is None:
                    yearThree = 0
                    total -= 1

                if total == 0:
                    avg = '-'
                else:
                    avg = (yearOne + yearTwo + yearThree) / total

                for column in range(2, max_col):
                    cell = sheet.cell(1, column)
                    if cell.value != name:
                        continue
                    sheet.cell(row=row, column=cell.column).value = yearOne
                    sheet.cell(row=row, column=cell.column + 1).value = yearTwo
                    sheet.cell(row=row, column=cell.column + 2).value = yearThree
                    sheet.cell(row=row, column=cell.column + 3).value = avg
    wb.save(path)
