import openpyxl
from Models.stock_exchange_information import StockExchangeInformation


def read_sp500_stock_exchange_information(path, sheet, ticker_column, exchange_column, row_start):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = wb[sheet]
    max_row = sheet.max_row
    data = []
    for i in range(row_start, max_row + 1):
        ticker_obj = sheet.cell(row=i, column=ticker_column)
        exchange_obj = sheet.cell(row=i, column=exchange_column)
        stock = StockExchangeInformation(ticker_obj.value, exchange_obj.value)
        data.append(stock)
    return data

